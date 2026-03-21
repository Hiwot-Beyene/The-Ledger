from __future__ import annotations

import asyncio
import csv
import os
import time
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Any, TypedDict

from langgraph.graph import END, StateGraph
from openpyxl import load_workbook

from ledger.agents.base_agent import BaseApexAgent
from ledger.schema.events import (
    CreditAnalysisRequested,
    DocumentAdded,
    DocumentFormat,
    DocumentFormatValidated,
    DocumentType,
    DocumentUploaded,
    ExtractionCompleted,
    ExtractionStarted,
    FinancialFacts,
    PackageCreated,
    PackageReadyForAnalysis,
    QualityAssessmentCompleted,
)


MAX_OCC_RETRIES = 5
PIPELINE_VERSION = "week3-compatible-1.0"
EXTRACTION_MODEL = "deterministic-parser-v1"

NUMERIC_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "total_revenue": ("total_revenue", "revenue", "sales"),
    "ebitda": ("ebitda",),
    "total_assets": ("total_assets", "assets"),
    "total_liabilities": ("total_liabilities", "liabilities"),
    "total_equity": ("total_equity", "equity"),
    "net_income": ("net_income", "net_profit"),
}


@dataclass(slots=True)
class UploadedDocument:
    document_id: str
    document_type: DocumentType
    document_format: DocumentFormat
    file_path: str
    filename: str
    uploaded_at: str


class DocumentProcessingState(TypedDict):
    application_id: str
    session_id: str
    package_id: str
    all_documents: list[UploadedDocument]
    processable_documents: list[UploadedDocument]
    extracted_facts_by_doc: dict[str, dict[str, Any]]
    quality_assessment: dict[str, Any] | None
    errors: list[str]
    output_events: list[dict[str, Any]]
    next_agent: str | None


class DocumentProcessingAgent(BaseApexAgent):
    def build_graph(self):
        g = StateGraph(DocumentProcessingState)
        g.add_node("validate_inputs", self._node_validate_inputs)
        g.add_node("open_aggregate_record", self._node_open_aggregate_record)
        g.add_node("load_external_data", self._node_load_external_data)
        g.add_node("process_documents", self._node_process_documents)
        g.add_node("assess_quality", self._node_assess_quality)
        g.add_node("write_output", self._node_write_output)
        g.set_entry_point("validate_inputs")
        g.add_edge("validate_inputs", "open_aggregate_record")
        g.add_edge("open_aggregate_record", "load_external_data")
        g.add_edge("load_external_data", "process_documents")
        g.add_edge("process_documents", "assess_quality")
        g.add_edge("assess_quality", "write_output")
        g.add_edge("write_output", END)
        return g.compile()

    def _initial_state(self, application_id: str) -> DocumentProcessingState:
        return DocumentProcessingState(
            application_id=application_id,
            session_id=self.session_id,
            package_id=f"pkg-{application_id}",
            all_documents=[],
            processable_documents=[],
            extracted_facts_by_doc={},
            quality_assessment=None,
            errors=[],
            output_events=[],
            next_agent=None,
        )

    async def _append_events_with_retry(
        self,
        stream_id: str,
        events: list[dict[str, Any]],
        *,
        causation_id: str | None = None,
        correlation_id: str | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> int:
        for attempt in range(MAX_OCC_RETRIES):
            try:
                ver = await self.store.stream_version(stream_id)
                return await self.store.append(
                    stream_id=stream_id,
                    events=events,
                    expected_version=ver,
                    causation_id=causation_id,
                    correlation_id=correlation_id,
                    metadata=metadata or {},
                )
            except Exception as exc:
                if "OptimisticConcurrencyError" in type(exc).__name__ and attempt < MAX_OCC_RETRIES - 1:
                    await asyncio.sleep(0.1 * (2**attempt))
                    continue
                raise

    async def _append_domain_error(
        self,
        stream_id: str,
        *,
        node_name: str,
        document_id: str | None,
        code: str,
        message: str,
        details: dict[str, Any] | None = None,
    ) -> None:
        event = {
            "event_type": "DomainError",
            "event_version": 1,
            "payload": {
                "application_id": self.application_id,
                "session_id": self.session_id,
                "node_name": node_name,
                "document_id": document_id,
                "code": code,
                "message": message,
                "details": details or {},
                "recorded_at": datetime.now().isoformat(),
            },
        }
        await self._append_events_with_retry(
            stream_id,
            [event],
            causation_id=self.session_id,
            correlation_id=self.application_id,
        )

    async def _node_validate_inputs(self, state: DocumentProcessingState) -> DocumentProcessingState:
        t0 = time.time()
        app_id = state["application_id"]
        loan_stream = f"loan-{app_id}"
        loan_events = await self.store.load_stream(loan_stream)
        await self._record_tool_call(
            "event_store.load_stream",
            f"stream_id={loan_stream}",
            f"events={len(loan_events)}",
            int((time.time() - t0) * 1000),
        )

        documents: list[UploadedDocument] = []
        for evt in loan_events:
            if evt["event_type"] != "DocumentUploaded":
                continue
            payload = evt["payload"]
            try:
                uploaded = DocumentUploaded(event_type="DocumentUploaded", **payload)
                documents.append(
                    UploadedDocument(
                        document_id=uploaded.document_id,
                        document_type=uploaded.document_type,
                        document_format=uploaded.document_format,
                        file_path=uploaded.file_path,
                        filename=uploaded.filename,
                        uploaded_at=uploaded.uploaded_at.isoformat(),
                    )
                )
            except Exception as exc:
                state["errors"].append(f"invalid DocumentUploaded payload: {exc}")

        required_types = {
            DocumentType.INCOME_STATEMENT,
            DocumentType.BALANCE_SHEET,
        }
        present_types = {doc.document_type for doc in documents}
        missing = [t.value for t in required_types if t not in present_types]
        if not documents or missing:
            errors = ["No uploaded documents found"] if not documents else []
            errors.extend([f"Missing required document type: {m}" for m in missing])
            await self._record_input_failed(
                missing_inputs=["document_uploads", *missing],
                validation_errors=errors,
            )
            await self._record_node_execution(
                "validate_inputs",
                ["application_id"],
                ["errors"],
                int((time.time() - t0) * 1000),
            )
            return {**state, "all_documents": documents, "errors": state["errors"] + errors}

        ms = int((time.time() - t0) * 1000)
        await self._record_input_validated(
            ["application_id", "DocumentUploaded", "required_document_types"],
            ms,
        )
        await self._record_node_execution(
            "validate_inputs",
            ["application_id"],
            ["all_documents"],
            ms,
        )
        return {**state, "all_documents": documents}

    async def _node_open_aggregate_record(self, state: DocumentProcessingState) -> DocumentProcessingState:
        t0 = time.time()
        app_id = state["application_id"]
        pkg_stream = f"docpkg-{app_id}"
        current = await self.store.load_stream(pkg_stream)
        await self._record_tool_call(
            "event_store.load_stream",
            f"stream_id={pkg_stream}",
            f"events={len(current)}",
            int((time.time() - t0) * 1000),
        )

        if not any(e["event_type"] == "PackageCreated" for e in current):
            created = PackageCreated(
                package_id=state["package_id"],
                application_id=app_id,
                required_documents=[DocumentType.INCOME_STATEMENT, DocumentType.BALANCE_SHEET],
                created_at=datetime.now(),
            ).to_store_dict()
            await self._append_events_with_retry(
                pkg_stream,
                [created],
                causation_id=self.session_id,
                correlation_id=app_id,
            )

        ms = int((time.time() - t0) * 1000)
        await self._record_node_execution(
            "open_aggregate_record",
            ["application_id", "package_id"],
            ["package_stream_ready"],
            ms,
        )
        return state

    async def _node_load_external_data(self, state: DocumentProcessingState) -> DocumentProcessingState:
        t0 = time.time()
        app_id = state["application_id"]
        pkg_stream = f"docpkg-{app_id}"
        existing = await self.store.load_stream(pkg_stream)
        await self._record_tool_call(
            "event_store.load_stream",
            f"stream_id={pkg_stream}",
            f"events={len(existing)}",
            int((time.time() - t0) * 1000),
        )

        prior_upload_requested = False
        all_events = []
        async for evt in self.store.load_all(event_types=["DocumentUploadRequested"]):
            all_events.append(evt)
            if evt["payload"].get("application_id") == app_id:
                prior_upload_requested = True
                break
        await self._record_tool_call(
            "event_store.load_all",
            "event_type=DocumentUploadRequested",
            f"matched={prior_upload_requested}",
            int((time.time() - t0) * 1000),
        )

        existing_doc_ids = {
            e["payload"].get("document_id")
            for e in existing
            if e["event_type"] == "DocumentAdded"
        }
        to_add: list[dict[str, Any]] = []
        for doc in state["all_documents"]:
            if doc.document_id in existing_doc_ids:
                continue
            to_add.append(
                DocumentAdded(
                    package_id=state["package_id"],
                    document_id=doc.document_id,
                    document_type=doc.document_type,
                    document_format=doc.document_format,
                    file_hash=self._sha(doc.file_path),
                    added_at=datetime.now(),
                ).to_store_dict()
            )
        if to_add:
            await self._append_events_with_retry(
                pkg_stream,
                to_add,
                causation_id=self.session_id,
                correlation_id=app_id,
            )

        errors = list(state["errors"])
        if not prior_upload_requested:
            msg = "DocumentUploadRequested not found for application"
            errors.append(msg)
            await self._append_domain_error(
                pkg_stream,
                node_name="load_external_data",
                document_id=None,
                code="missing_prerequisite",
                message=msg,
                details={"application_id": app_id},
            )

        ms = int((time.time() - t0) * 1000)
        await self._record_node_execution(
            "load_external_data",
            ["all_documents"],
            ["processable_documents", "errors"],
            ms,
        )
        return {**state, "processable_documents": state["all_documents"], "errors": errors}

    async def _node_process_documents(self, state: DocumentProcessingState) -> DocumentProcessingState:
        t0 = time.time()
        app_id = state["application_id"]
        pkg_stream = f"docpkg-{app_id}"
        docs = state["processable_documents"]
        extracted: dict[str, dict[str, Any]] = {}
        errors = list(state["errors"])

        for doc in docs:
            val_start = time.time()
            valid, reason, page_count, detected_fmt = self._validate_file(doc.file_path, doc.document_format)
            validated = DocumentFormatValidated(
                package_id=state["package_id"],
                document_id=doc.document_id,
                document_type=doc.document_type,
                page_count=page_count,
                detected_format=detected_fmt,
                validated_at=datetime.now(),
            ).to_store_dict()
            await self._append_events_with_retry(
                pkg_stream,
                [validated],
                causation_id=self.session_id,
                correlation_id=app_id,
                metadata={"file_path": doc.file_path, "validation_ok": valid, "reason": reason},
            )
            await self._record_tool_call(
                "filesystem.validate_document",
                f"path={doc.file_path}",
                f"valid={valid} format={detected_fmt}",
                int((time.time() - val_start) * 1000),
            )
            if not valid:
                err = f"{doc.document_id}: {reason}"
                errors.append(err)
                await self._append_domain_error(
                    pkg_stream,
                    node_name="process_documents",
                    document_id=doc.document_id,
                    code="invalid_document_format",
                    message=reason,
                    details={"file_path": doc.file_path},
                )
                continue

            started = ExtractionStarted(
                package_id=state["package_id"],
                document_id=doc.document_id,
                document_type=doc.document_type,
                pipeline_version=PIPELINE_VERSION,
                extraction_model=EXTRACTION_MODEL,
                started_at=datetime.now(),
            ).to_store_dict()
            await self._append_events_with_retry(
                pkg_stream,
                [started],
                causation_id=self.session_id,
                correlation_id=app_id,
                metadata={"file_path": doc.file_path},
            )

            extract_start = time.time()
            try:
                facts = await self._extract_facts(doc)
            except Exception as exc:
                errors.append(f"{doc.document_id}: extraction failed: {exc}")
                await self._append_domain_error(
                    pkg_stream,
                    node_name="process_documents",
                    document_id=doc.document_id,
                    code="extraction_failure",
                    message=str(exc),
                    details={"file_path": doc.file_path},
                )
                facts = FinancialFacts(
                    extraction_notes=[
                        "extraction_failed",
                        f"source_file={doc.file_path}",
                    ],
                    field_confidence={k: 0.0 for k in NUMERIC_FIELD_ALIASES.keys()},
                )

            completed = ExtractionCompleted(
                package_id=state["package_id"],
                document_id=doc.document_id,
                document_type=doc.document_type,
                facts=facts,
                raw_text_length=0,
                tables_extracted=1,
                processing_ms=int((time.time() - extract_start) * 1000),
                completed_at=datetime.now(),
            ).to_store_dict()
            await self._append_events_with_retry(
                pkg_stream,
                [completed],
                causation_id=self.session_id,
                correlation_id=app_id,
                metadata={"file_path": doc.file_path},
            )
            await self._record_tool_call(
                "document_extraction",
                f"doc_id={doc.document_id} format={doc.document_format.value}",
                "ExtractionCompleted",
                int((time.time() - extract_start) * 1000),
            )
            extracted[doc.document_id] = completed["payload"]["facts"] or {}

        ms = int((time.time() - t0) * 1000)
        await self._record_node_execution(
            "process_documents",
            ["processable_documents"],
            ["extracted_facts_by_doc", "errors"],
            ms,
        )
        return {**state, "extracted_facts_by_doc": extracted, "errors": errors}

    async def _node_assess_quality(self, state: DocumentProcessingState) -> DocumentProcessingState:
        t0 = time.time()
        app_id = state["application_id"]
        pkg_stream = f"docpkg-{app_id}"

        merged = self._merge_facts(list(state["extracted_facts_by_doc"].values()))
        quality = self._deterministic_quality_assessment(merged)

        quality_event = QualityAssessmentCompleted(
            package_id=state["package_id"],
            document_id="package-summary",
            overall_confidence=quality["overall_confidence"],
            is_coherent=quality["is_coherent"],
            anomalies=quality["anomalies"],
            critical_missing_fields=quality["critical_missing_fields"],
            reextraction_recommended=quality["reextraction_recommended"],
            auditor_notes=quality["auditor_notes"],
            assessed_at=datetime.now(),
        ).to_store_dict()
        await self._append_events_with_retry(
            pkg_stream,
            [quality_event],
            causation_id=self.session_id,
            correlation_id=app_id,
        )

        ms = int((time.time() - t0) * 1000)
        await self._record_node_execution(
            "assess_quality",
            ["extracted_facts_by_doc"],
            ["quality_assessment"],
            ms,
        )
        return {**state, "quality_assessment": quality}

    async def _node_write_output(self, state: DocumentProcessingState) -> DocumentProcessingState:
        t0 = time.time()
        app_id = state["application_id"]
        pkg_stream = f"docpkg-{app_id}"
        loan_stream = f"loan-{app_id}"
        quality = state.get("quality_assessment") or {}

        ready_event = PackageReadyForAnalysis(
            package_id=state["package_id"],
            application_id=app_id,
            documents_processed=len(state["extracted_facts_by_doc"]),
            has_quality_flags=bool(quality.get("anomalies") or quality.get("critical_missing_fields")),
            quality_flag_count=len(quality.get("anomalies", [])) + len(quality.get("critical_missing_fields", [])),
            ready_at=datetime.now(),
        ).to_store_dict()
        await self._append_events_with_retry(
            pkg_stream,
            [ready_event],
            causation_id=self.session_id,
            correlation_id=app_id,
        )

        trigger_event = CreditAnalysisRequested(
            application_id=app_id,
            requested_at=datetime.now(),
            requested_by=self.agent_id,
            priority="NORMAL",
        ).to_store_dict()
        await self._append_events_with_retry(
            loan_stream,
            [trigger_event],
            causation_id=self.session_id,
            correlation_id=app_id,
        )

        events_written = [
            {"stream_id": pkg_stream, "event_type": "PackageReadyForAnalysis"},
            {"stream_id": loan_stream, "event_type": "CreditAnalysisRequested"},
        ]
        await self._record_output_written(
            events_written,
            f"Processed {len(state['extracted_facts_by_doc'])} documents; quality coherent={quality.get('is_coherent')}.",
        )

        ms = int((time.time() - t0) * 1000)
        await self._record_node_execution(
            "write_output",
            ["quality_assessment"],
            ["output_events", "next_agent"],
            ms,
        )
        return {**state, "output_events": events_written, "next_agent": "credit_analysis"}

    async def _extract_facts(self, doc: UploadedDocument) -> FinancialFacts:
        if doc.document_format == DocumentFormat.CSV:
            return self._extract_from_csv(doc.file_path)
        if doc.document_format == DocumentFormat.XLSX:
            return self._extract_from_xlsx(doc.file_path)
        if doc.document_format == DocumentFormat.PDF:
            return await self._extract_from_pdf(doc.file_path, doc.document_type)
        raise ValueError(f"Unsupported format: {doc.document_format.value}")

    async def _extract_from_pdf(self, file_path: str, document_type: DocumentType) -> FinancialFacts:
        try:
            from document_refinery.pipeline import extract_financial_facts  # type: ignore

            out = extract_financial_facts(file_path, document_type.value)
            raw = await out if asyncio.iscoroutine(out) else out
            return FinancialFacts(**(raw or {}))
        except Exception:
            return FinancialFacts(
                extraction_notes=[
                    "week3_pipeline_unavailable_or_failed",
                    f"source_file={file_path}",
                ],
                field_confidence={k: 0.0 for k in NUMERIC_FIELD_ALIASES.keys()},
            )

    def _extract_from_csv(self, file_path: str) -> FinancialFacts:
        values: dict[str, Decimal] = {}
        with open(file_path, "r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
        for row in rows:
            for target, aliases in NUMERIC_FIELD_ALIASES.items():
                for alias in aliases:
                    if alias not in row or row[alias] in (None, ""):
                        continue
                    num = self._to_decimal(row[alias])
                    if num is not None:
                        values[target] = num
        return FinancialFacts(
            **values,
            extraction_notes=[f"source_file={file_path}"],
            field_confidence={k: (1.0 if k in values else 0.0) for k in NUMERIC_FIELD_ALIASES.keys()},
        )

    def _extract_from_xlsx(self, file_path: str) -> FinancialFacts:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        values: dict[str, Decimal] = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                if not row or row[0] is None:
                    continue
                key = str(row[0]).strip().lower().replace(" ", "_")
                value = self._to_decimal(row[1])
                if value is None:
                    continue
                for target, aliases in NUMERIC_FIELD_ALIASES.items():
                    if key in aliases:
                        values[target] = value
        wb.close()
        return FinancialFacts(
            **values,
            extraction_notes=[f"source_file={file_path}"],
            field_confidence={k: (1.0 if k in values else 0.0) for k in NUMERIC_FIELD_ALIASES.keys()},
        )

    @staticmethod
    def _to_decimal(value: Any) -> Decimal | None:
        if value is None:
            return None
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        if text.startswith("$"):
            text = text[1:]
        try:
            return Decimal(text)
        except Exception:
            return None

    def _merge_facts(self, facts_list: list[dict[str, Any]]) -> dict[str, Any]:
        merged: dict[str, Any] = {}
        for facts in facts_list:
            for k, v in facts.items():
                if v is not None and k not in merged:
                    merged[k] = v
        for required in ("total_revenue", "ebitda", "total_assets", "total_liabilities"):
            merged.setdefault(required, None)
        return merged

    def _deterministic_quality_assessment(self, facts: dict[str, Any]) -> dict[str, Any]:
        critical = ["total_revenue", "net_income", "total_assets", "total_liabilities"]
        missing = [k for k in critical if facts.get(k) in (None, "")]
        anomalies: list[str] = []

        total_assets = self._to_decimal(facts.get("total_assets"))
        total_liabilities = self._to_decimal(facts.get("total_liabilities"))
        total_equity = self._to_decimal(facts.get("total_equity"))
        if total_assets is not None and total_liabilities is not None and total_equity is not None:
            diff = total_assets - (total_liabilities + total_equity)
            if abs(diff) > Decimal("1.00"):
                anomalies.append(f"balance_sheet_not_balanced:{diff}")

        total_revenue = self._to_decimal(facts.get("total_revenue"))
        ebitda = self._to_decimal(facts.get("ebitda"))
        if total_revenue and ebitda:
            margin = float(ebitda / total_revenue) if total_revenue != 0 else 0.0
            if margin > 0.80 or margin < -0.40:
                anomalies.append(f"implausible_ebitda_margin:{margin:.2f}")

        confidence = max(0.0, min(1.0, 1.0 - (0.15 * len(missing)) - (0.1 * len(anomalies))))
        return {
            "overall_confidence": round(confidence, 3),
            "is_coherent": not anomalies and not missing,
            "anomalies": anomalies,
            "critical_missing_fields": missing,
            "reextraction_recommended": bool(missing or anomalies),
            "auditor_notes": "Deterministic quality checks over normalized GAAP fields.",
        }

    def _validate_file(
        self, file_path: str, declared_format: DocumentFormat
    ) -> tuple[bool, str, int, str]:
        if not os.path.exists(file_path):
            return False, "file_not_found", 0, declared_format.value
        if os.path.getsize(file_path) == 0:
            return False, "empty_file", 0, declared_format.value

        ext = os.path.splitext(file_path)[1].lower().lstrip(".")
        detected_format = ext or declared_format.value
        if detected_format != declared_format.value:
            return False, f"format_mismatch: declared={declared_format.value} detected={detected_format}", 0, detected_format

        if declared_format == DocumentFormat.PDF:
            with open(file_path, "rb") as handle:
                data = handle.read(4096)
            if not data.startswith(b"%PDF"):
                return False, "invalid_pdf_header", 0, detected_format
            page_count = max(1, data.count(b"/Type /Page"))
            return True, "ok", page_count, detected_format
        if declared_format == DocumentFormat.CSV:
            with open(file_path, "r", encoding="utf-8", newline="") as handle:
                reader = csv.reader(handle)
                first = next(reader, None)
            if not first:
                return False, "csv_missing_header", 0, detected_format
            return True, "ok", 1, detected_format
        if declared_format == DocumentFormat.XLSX:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = len(wb.sheetnames)
            wb.close()
            if sheets == 0:
                return False, "xlsx_no_sheets", 0, detected_format
            return True, "ok", sheets, detected_format
        return False, "unsupported_format", 0, detected_format
